# -*- coding: utf-8 -*-
"""
Excel export module for Recruitment Agent.
Provides: export_candidates(), export_interviews()
Returns bytes (for Streamlit download_button).
"""

import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


def _style_header(ws, num_cols):
    """Apply header styling."""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def _auto_width(ws):
    """Auto-fit column widths."""
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def _add_filters(ws, num_cols):
    """Add auto-filter to header row."""
    last_col = get_column_letter(num_cols)
    ws.auto_filter.ref = f"A1:{last_col}1"


def _freeze_header(ws):
    """Freeze the header row."""
    ws.freeze_panes = "A2"


def export_candidates(candidates):
    """
    Export candidate list to Excel bytes.
    Args:
        candidates: list of dict (from db.get_candidates())
    Returns:
        bytes of .xlsx file
    """
    rows = []
    for c in candidates:
        rows.append({
            "ID": c.get("id", ""),
            "Name": c.get("name", ""),
            "Phone": c.get("phone", ""),
            "Email": c.get("email", ""),
            "Position": c.get("recent_position", ""),
            "Company": c.get("recent_company", ""),
            "Status": c.get("status", ""),
            "Source": c.get("source", ""),
            "Match Score": c.get("match_score", ""),
            "Created At": c.get("created_at", ""),
            "Notes": c.get("notes", ""),
        })

    df = pd.DataFrame(rows)

    # Write to BytesIO using openpyxl for styling
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidates"

    # Headers
    headers = list(df.columns)
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)

    # Data
    for ri, row_data in enumerate(df.values, 2):
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            # Zebra striping for readability
            if ri % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    _style_header(ws, len(headers))
    if len(headers) > 0:
        _add_filters(ws, len(headers))
    _freeze_header(ws)
    _auto_width(ws)
    
    # Add summary sheet
    ws_summary = wb.create_sheet("Summary")
    ws_summary["A1"] = "Export Summary"
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary["A3"] = "Total Candidates:"
    ws_summary["B3"] = len(candidates)
    ws_summary["A4"] = "Export Time:"
    ws_summary["B4"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def export_interviews(interviews):
    """
    Export interview records to Excel bytes.
    Args:
        interviews: list of dict (from db.get_interviews())
    Returns:
        bytes of .xlsx file
    """
    rows = []
    for iv in interviews:
        rows.append({
            "ID": iv.get("id", ""),
            "Candidate Name": iv.get("candidate_name", ""),
            "Job Title": iv.get("job_title", ""),
            "Interview Type": iv.get("interview_type", ""),
            "Scheduled Time": iv.get("scheduled_time", ""),
            "Status": iv.get("status", ""),
            "Interviewer": iv.get("interviewer", ""),
            "Feedback": iv.get("feedback", ""),
            "Score": iv.get("score", ""),
            "Created At": iv.get("created_at", ""),
        })

    df = pd.DataFrame(rows)

    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Interviews"

    headers = list(df.columns)
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)

    for ri, row_data in enumerate(df.values, 2):
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            # Zebra striping
            if ri % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    _style_header(ws, len(headers))
    if len(headers) > 0:
        _add_filters(ws, len(headers))
    _freeze_header(ws)
    _auto_width(ws)
    
    # Add summary sheet
    ws_summary = wb.create_sheet("Summary")
    ws_summary["A1"] = "Export Summary"
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary["A3"] = "Total Interviews:"
    ws_summary["B3"] = len(interviews)
    ws_summary["A4"] = "Export Time:"
    ws_summary["B4"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Calculate statistics
    ws_summary["A6"] = "Statistics:"
    ws_summary["A6"].font = Font(bold=True)
    
    # Count by status
    status_counts = {}
    for iv in interviews:
        status = iv.get("status", "Unknown")
        status_counts[status] = status_counts.get(status, 0) + 1
    
    row = 7
    for status, count in status_counts.items():
        ws_summary[f"A{row}"] = status
        ws_summary[f"B{row}"] = count
        row += 1
    
    wb.save(output)
    output.seek(0)
    return output.getvalue()
